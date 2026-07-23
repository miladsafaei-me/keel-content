"""``./manage.py contentplan_ingest_youtube <url> [<url> ...]`` — the YouTube-transcript
intake route into the ContentPlan roadmap (path 4 for blogs).

A YouTube link (given one-off in chat, or as a column in an Excel/CSV/TXT list) is
turned into a blog the same way every other route is: this command extracts the
video's transcript with ``yt-dlp`` (zero cost, no API key — see
``keel_content.core.youtube_transcript``), attaches it to a ``source_type=youtube``
ContentPlan row via the shared upsert core, and stores the source video URL. From
there the row flows through the SAME machinery as every other source — the reconcile
cannibalization gate (keyed on ``canonical_key``), cluster placement, the brief stage,
generation (the transcript is the PRIMARY source the article is written from), and
``content_import`` (which mirrors the video URL onto the produced Post). So a
YouTube-sourced blog cannot cannibalize a keyword/top-pages blog, and it takes its
place in the intent registry + a topic cluster like any other page.

Two ways to run it:

1. Extract only (get the transcript + a spec scaffold to classify, no DB write)::

     ./manage.py contentplan_ingest_youtube <url> --extract-only --out /tmp/yt.json

   Emits ``[{youtube_url, video_id, title, channel, duration, source_transcript,
   intent, topic_cluster, role, markets, ...}]`` — the classification fields are left
   blank for a human/LLM to fill (intent, cluster, market, canonical_key), exactly
   like ``/seo-clustering`` emits a spec for review before ingest.

2. Ingest a classified spec list (or bare URLs for a minimally-classified row)::

     ./manage.py contentplan_ingest_youtube --json /tmp/yt.json
     ./manage.py contentplan_ingest_youtube https://youtu.be/<id> --cluster "ICT Trading"

Input precedence: ``--json`` specs, then ``--file`` rows, then positional URLs.
Each spec's transcript is extracted here unless it already carries ``source_transcript``
(or a reachable ``source_transcript_path``, which is read in). Bare URLs create a
``planned`` row titled from the video; enrich its intent/cluster before generation.

The classification (intent / cluster / market / canonical_key / title) is normally
produced automatically by ``tools/content_pipeline/classify_youtube.workflow.js`` — one
agent per video reads the transcript + the live taxonomy and emits a ``--json`` spec that
joins the best-fit cluster or proposes a new one — the same shape ``/seo-clustering``
produces for the keyword route.
"""

from __future__ import annotations

import csv
import json
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.utils.text import slugify

from keel_content import host
from keel_content.adapters import get_adapter

# Resolve the configured publisher adapter (default: the reference SignalBots adapter).
upsert_content_plan_spec = get_adapter().upsert_content_plan_spec
from keel_content.core import youtube_transcript as yt

ContentPlan = host.content_plan_model()

# Optional per-row classification columns an Excel/CSV list may carry. Comma-split
# ones become lists; the rest are scalar strings copied straight onto the spec.
_LIST_COLUMNS = ("markets", "categories", "audience_roles", "audience_levels", "glossary_terms")
_SCALAR_COLUMNS = ("title", "h1", "intent", "intent_frame", "entity", "role", "topic_cluster",
                   "topic_cluster_slug", "canonical_key", "slug")
_URL_HEADERS = ("youtube_url", "url", "video", "video_url", "link", "youtube")


class Command(BaseCommand):
    help = "Ingest YouTube video(s) as ContentPlan rows written from their transcript (path 4)."

    def add_arguments(self, parser):
        parser.add_argument("urls", nargs="*", help="One or more YouTube URLs.")
        parser.add_argument("--json", dest="json_path", help="A JSON list of specs (each with a youtube_url).")
        parser.add_argument("--file", dest="file_path", help="An .xlsx/.csv/.txt list of video URLs (+ optional columns).")
        parser.add_argument("--cluster", help="Topic cluster name to place ALL ingested videos into (override).")
        parser.add_argument("--lang", default="en", help="Caption language to prefer (default: en).")
        parser.add_argument("--extract-only", action="store_true",
                            help="Extract transcripts + write a spec scaffold; do NOT touch the DB.")
        parser.add_argument("--out", help="Where --extract-only writes the scaffold JSON (default: stdout).")
        parser.add_argument("--replace", action="store_true",
                            help="Overwrite rows already in production (generating/drafted/published).")
        parser.add_argument("--dry-run", action="store_true")

    def handle(self, *args, **opts):
        specs = self._collect_specs(opts)
        if not specs:
            raise CommandError("no videos given — pass URLs, --file, or --json")

        cluster_override = (opts.get("cluster") or "").strip()
        lang = opts["lang"]
        scaffold: list[dict] = []
        created = updated = skipped = locked = failed = 0

        for spec in specs:
            url = (spec.get("youtube_url") or spec.get("url") or "").strip()
            if not yt.video_id(url):
                self.stderr.write(self.style.WARNING(f"  - skip (not a YouTube URL): {url!r}"))
                skipped += 1
                continue

            # A classified spec may point at the transcript by path (from the classify
            # workflow) instead of inlining it; read it in when the file is reachable.
            tpath = (spec.get("source_transcript_path") or "").strip()
            if tpath and not (spec.get("source_transcript") or "").strip():
                tp = Path(tpath).expanduser()
                if tp.is_file():
                    spec["source_transcript"] = tp.read_text(encoding="utf-8")

            # Extract the transcript unless the spec already carries one.
            if not (spec.get("source_transcript") or "").strip():
                try:
                    meta = yt.extract(url, lang=lang)
                except yt.TranscriptUnavailable as exc:
                    self.stderr.write(self.style.ERROR(f"  ! fail {url}: {exc}"))
                    failed += 1
                    continue
                spec.setdefault("title", meta["title"])
                spec["youtube_url"] = meta["url"]
                spec["source_transcript"] = meta["transcript"]
                spec["_video_id"] = meta["video_id"]
                spec.setdefault("_channel", meta["channel"])
                spec.setdefault("_duration", meta["duration"])
                self.stdout.write(f"  ~ transcript: {meta['title'][:56]} ({meta['words']} words)")
            spec.setdefault("_video_id", yt.video_id(url))

            if cluster_override:
                spec["topic_cluster"] = cluster_override
            spec.setdefault("role", "spoke")
            spec.setdefault("target", "blog")
            if not (spec.get("slug") or "").strip():
                spec["slug"] = slugify(spec.get("title") or spec["_video_id"])[:255]

            if opts["extract_only"]:
                scaffold.append(self._scaffold_row(spec))
                continue

            if opts["dry_run"]:
                self.stdout.write(f"    would ingest {spec['slug']}")
                created += 1
                continue

            plan, outcome = upsert_content_plan_spec(
                spec,
                source_type=ContentPlan.Source.YOUTUBE.value,
                source_ref=f"youtube:{spec['_video_id']}",
                replace=opts["replace"],
            )
            slug = plan.slug if plan else spec.get("slug", "?")
            if outcome == "created":
                created += 1
                self.stdout.write(self.style.SUCCESS(f"    + create {slug}"))
            elif outcome == "updated":
                updated += 1
                self.stdout.write(self.style.SUCCESS(f"    * update {slug}"))
            elif outcome == "locked":
                locked += 1
                self.stdout.write(f"    = lock   {slug} (in production; --replace to overwrite)")
            else:
                skipped += 1
                self.stdout.write(f"    - skip   {slug} ({outcome})")

        if opts["extract_only"]:
            self._emit_scaffold(scaffold, opts.get("out"))
            self.stdout.write(self.style.SUCCESS(
                f"\nextracted {len(scaffold)} transcript(s), {failed} failed. "
                "Fill intent / topic_cluster / markets / canonical_key, then ingest with --json."
            ))
            return

        self.stdout.write("")
        self.stdout.write(self.style.SUCCESS(
            f"done: {created} created, {updated} updated, {locked} locked, "
            f"{skipped} skipped, {failed} failed (of {len(specs)})"
        ))
        if (created or updated) and not opts["dry_run"]:
            self.stdout.write(
                "  rows are status=planned — they must pass the reconcile gate "
                "(cannibalization) before export_worklist will generate them."
            )

    # Input collection: merge --json specs, --file rows, and positional URLs.
    def _collect_specs(self, opts) -> list[dict]:
        specs: list[dict] = []
        if opts.get("json_path"):
            path = Path(opts["json_path"]).expanduser()
            if not path.is_file():
                raise CommandError(f"--json file not found: {path}")
            data = json.loads(path.read_text(encoding="utf-8"))
            if isinstance(data, dict):
                data = data.get("contents") or data.get("videos") or data.get("specs") or []
            if not isinstance(data, list):
                raise CommandError("--json must be a list of spec objects")
            specs.extend(d for d in data if isinstance(d, dict))
        if opts.get("file_path"):
            specs.extend(self._read_file(Path(opts["file_path"]).expanduser()))
        for url in opts.get("urls") or []:
            specs.append({"youtube_url": url})
        return specs

    def _read_file(self, path: Path) -> list[dict]:
        if not path.is_file():
            raise CommandError(f"--file not found: {path}")
        suffix = path.suffix.lower()
        if suffix in (".txt", ""):
            return [
                {"youtube_url": line.strip()}
                for line in path.read_text(encoding="utf-8").splitlines()
                if line.strip() and not line.strip().startswith("#")
            ]
        if suffix == ".csv":
            with path.open(encoding="utf-8", newline="") as fh:
                return self._rows_to_specs(list(csv.reader(fh)))
        if suffix in (".xlsx", ".xlsm"):
            try:
                from openpyxl import load_workbook
            except ImportError as exc:  # pragma: no cover - depends on env
                raise CommandError("reading .xlsx needs openpyxl (pip install openpyxl)") from exc
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            rows = [[("" if c is None else str(c)) for c in row] for row in ws.iter_rows(values_only=True)]
            return self._rows_to_specs(rows)
        raise CommandError(f"unsupported --file type: {suffix} (use .xlsx/.csv/.txt)")

    def _rows_to_specs(self, rows: list[list[str]]) -> list[dict]:
        rows = [r for r in rows if any((c or "").strip() for c in r)]
        if not rows:
            return []
        header = [str(c or "").strip().lower() for c in rows[0]]
        has_header = any(h in _URL_HEADERS for h in header) or any(
            h in (_SCALAR_COLUMNS + _LIST_COLUMNS) for h in header
        )
        if not has_header:
            # No header row — treat every first cell as a URL.
            return [{"youtube_url": r[0].strip()} for r in rows if r and r[0].strip()]

        url_idx = next((i for i, h in enumerate(header) if h in _URL_HEADERS), 0)
        col = {h: i for i, h in enumerate(header)}
        specs: list[dict] = []
        for r in rows[1:]:
            url = (r[url_idx] if url_idx < len(r) else "").strip()
            if not url:
                continue
            spec: dict = {"youtube_url": url}
            for name in _SCALAR_COLUMNS:
                i = col.get(name)
                if i is not None and i < len(r) and r[i].strip():
                    spec[name] = r[i].strip()
            for name in _LIST_COLUMNS:
                i = col.get(name)
                if i is not None and i < len(r) and r[i].strip():
                    spec[name] = [p.strip() for p in r[i].split(",") if p.strip()]
            specs.append(spec)
        return specs

    # Extract-only scaffold: one classification-ready row per video.
    def _scaffold_row(self, spec: dict) -> dict:
        return {
            "youtube_url": spec.get("youtube_url", ""),
            "video_id": spec.get("_video_id", ""),
            "title": spec.get("title", ""),
            "channel": spec.get("_channel", ""),
            "duration": spec.get("_duration", ""),
            "slug": spec.get("slug", ""),
            # Fill these in before ingest (like a clustering spec):
            "intent": spec.get("intent", ""),
            "intent_frame": spec.get("intent_frame", ""),
            "entity": spec.get("entity", ""),
            "role": spec.get("role", "spoke"),
            "topic_cluster": spec.get("topic_cluster", ""),
            "canonical_key": spec.get("canonical_key", ""),
            "markets": spec.get("markets", []),
            "categories": spec.get("categories", []),
            "audience_roles": spec.get("audience_roles", []),
            "audience_levels": spec.get("audience_levels", []),
            "glossary_terms": spec.get("glossary_terms", []),
            "target": "blog",
            "source_transcript": spec.get("source_transcript", ""),
        }

    def _emit_scaffold(self, scaffold: list[dict], out: str | None) -> None:
        payload = json.dumps(scaffold, ensure_ascii=False, indent=2)
        if out:
            Path(out).expanduser().write_text(payload + "\n", encoding="utf-8")
            self.stdout.write(f"  scaffold written to {out}")
        else:
            self.stdout.write(payload)
